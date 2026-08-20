import os
import csv
from datetime import datetime
import requests
import streamlit as st
import cv2
import streamlit.components.v1 as components
from textwrap import dedent
import matplotlib.pyplot as plt
import pandas as pd
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter
from modelo_ia import clasificar_imagen
from gemini_analisis import analizar_causas
from streamlit_option_menu import option_menu
import io
import re
from datetime import datetime
from html import escape
import base64
from PIL import Image
from pathlib import Path
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)


def cargar_tema(nombre_archivo):
    ruta = Path(__file__).parent / "styles" / nombre_archivo

    if ruta.exists():
        with open(ruta, "r", encoding="utf-8") as archivo:
            st.markdown(
                f"<style>{archivo.read()}</style>",
                unsafe_allow_html=True,
            )
    else:
        st.warning(f"No se encontró el archivo de tema: {ruta}")


def cargar_css(nombre_archivo):
    ruta = Path(__file__).parent / "assets" / "css" / nombre_archivo

    if ruta.exists():
        with open(ruta, "r", encoding="utf-8") as archivo:
            st.markdown(
                f"<style>{archivo.read()}</style>",
                unsafe_allow_html=True,
            )
    else:
        st.warning(f"No se encontró el archivo CSS: {ruta}")


# ---------------- APP PRINCIPAL ----------------

st.set_page_config(page_title="VisionQA", page_icon="🔍", layout="wide")
with open("styles/styles.css", encoding="utf-8") as f:
    st.markdown(
        f"<style>{f.read()}</style>",
        unsafe_allow_html=True,
    )
st.markdown(
    """
    <style>

    /* Fondo general */
    .stApp {
        background-color: #f4f7fb;
    }

    /* Ocultar elementos predeterminados de Streamlit */
    #MainMenu {
        visibility: hidden;
    }

    footer {
        visibility: hidden;
    }

    header {
        visibility: hidden;
    }

    /* Espacio del contenido principal */
    .block-container {
        padding-top: 1rem;
        padding-bottom: 2rem;
        max-width: 100%;
    }

    /* Barra superior del Dashboard */
    .visionqa-header {
    background: linear-gradient(90deg, #168db5, #24a0c1);
    border-radius: 10px;
    padding: 18px 24px;
    margin-bottom: 22px;
    color: white;
    box-shadow: 0 4px 14px rgba(0, 0, 0, 0.12);
    display: flex;
    justify-content: space-between;
    align-items: center;
}

.visionqa-header-left {
    display: flex;
    flex-direction: column;
}

.visionqa-header-right {
    text-align: right;
    color: white;
}

.visionqa-saludo {
    color: white;
    font-size: 20px;
    font-weight: 700;
}

.visionqa-subtitulo {
    color: white;
    font-size: 13px;
    margin-top: 5px;
    opacity: 0.92;
}

.visionqa-fecha {
    color: white;
    font-size: 13px;
}

.visionqa-hora {
    color: white;
    font-size: 18px;
    font-weight: 700;
    margin-top: 2px;
}
    </style>
    """,
    unsafe_allow_html=True,
)
# ---------------- ESTILOS ----------------

st.markdown(
    """
    <link
        rel="stylesheet"
        href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css"
    >
    """,
    unsafe_allow_html=True,
)
# ---------------- VARIABLES GLOBALES ----------------

archivo_csv = "registro_inspecciones.csv"

os.makedirs("inspecciones", exist_ok=True)

if "inspeccion" not in st.session_state:
    st.session_state.inspeccion = False


# ---------------- FUNCIONES ----------------
def guardar_inspeccion_api(resultado, defecto, confianza, archivo, origen):
    url_api = "http://127.0.0.1:8000/api/inspecciones/"

    datos = {
        "resultado": resultado,
        "defecto": defecto,
        "confianza": float(confianza),
        "archivo": archivo,
        "origen": origen,
    }

    try:
        respuesta = requests.post(url_api, json=datos, timeout=10)

        if respuesta.status_code == 201:
            return True, "Inspección guardada en Django."

        return False, f"Error de API {respuesta.status_code}: {respuesta.text}"

    except requests.exceptions.ConnectionError:
        return False, "No se pudo conectar con Django."

    except requests.exceptions.RequestException as error:
        return False, f"Error al enviar la inspección: {error}"


def obtener_inspecciones_api():

    url_api = "http://127.0.0.1:8000/api/inspecciones/"

    try:

        respuesta = requests.get(url_api, timeout=10)

        if respuesta.status_code == 200:

            return respuesta.json()

    except requests.exceptions.RequestException:

        pass

    return []


def mostrar_titulo(icono, titulo, descripcion):
    html_titulo = (
        '<div style="margin-top:30px; margin-bottom:22px;">'
        '<div style="'
        "display:flex;"
        "align-items:center;"
        "gap:10px;"
        "color:#231F20;"
        "font-size:40px;"
        "font-weight:700;"
        '">'
        f"{icono_svg(icono, 30, 0)}"
        f"<span>{titulo}</span>"
        "</div>"
        '<div style="'
        "margin-top:6px;"
        "color:#64748B;"
        "font-size:18px;"
        '">'
        f"{descripcion}"
        "</div>"
        "</div>"
    )

    st.markdown(
        html_titulo,
        unsafe_allow_html=True,
    )


def mostrar_encabezado_seccion(titulo, descripcion=""):

    html = (
        '<div class="section-header">'
        f'<div class="section-title">{titulo}</div>'
        f'<div class="section-description">{descripcion}</div>'
        "</div>"
    )

    st.markdown(html, unsafe_allow_html=True)


def mostrar_estado_sistema():
    mostrar_encabezado_seccion(
    "🟢 Estado del Sistema"
)

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown(
            """
            <div class="system-card">
                <div class="system-card-title"><i class="bi bi-cpu"></i> Modelo IA</div>
                <div class="system-card-text"><strong>Estado:</strong> Conectado</div>
                <div class="system-card-status">🟢 Cargado correctamente</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col2:
        st.markdown(
            """
            <div class="system-card">
                <div class="system-card-title"><i class="bi bi-stars"></i> Gemini</div>
                <div class="system-card-text"><strong>Estado:</strong> Disponible</div>
                <div class="system-card-status">🟢 Conectado</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col3:
        st.markdown(
            """
            <div class="system-card">
                <div class="system-card-title"><i class="bi bi-database"></i> Base de datos</div>
                <div class="system-card-text"><strong>Estado:</strong> Disponible</div>
                <div class="system-card-status">🟢 Registro listo</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.divider()

def cargar_datos_registro():

    url_api = "http://127.0.0.1:8000/api/inspecciones/"

    try:

        respuesta = requests.get(url_api, timeout=10)

        if respuesta.status_code != 200:

            return 0, 0, 0

        registros = respuesta.json()

        total = len(registros)
        aptas = 0
        no_aptas = 0

        for registro in registros:

            resultado = str(registro.get("resultado", "")).strip().upper()

            if resultado in ["APTO", "BUENA"]:

                aptas += 1

            elif resultado in ["NO APTO", "MALA"]:

                no_aptas += 1

        return total, aptas, no_aptas

    except requests.exceptions.RequestException:

        return 0, 0, 0


def mostrar_modulo_inspeccion():

    mostrar_encabezado_seccion(
        "Método de inspección",
        "Selecciona cómo deseas capturar la pieza para iniciar el análisis.",
    )

    if "metodo_inspeccion" not in st.session_state:
        st.session_state["metodo_inspeccion"] = None

    col_metodo_1, col_metodo_2, col_metodo_3 = st.columns(3)

    with col_metodo_1:
        carga_activa = st.session_state["metodo_inspeccion"] == "Cargar imagen"

        st.markdown(
            dedent(f"""
                <div class="inspection-card {'active' if carga_activa else ''}">
                    <div class="inspection-card-top"></div>
                    <div class="inspection-card-content">
                        <div class="inspection-card-icon"><i class="bi bi-image"></i></div>
                        <div class="inspection-card-title">Cargar imagen</div>
                        <div class="inspection-card-description">
                            Selecciona una imagen desde tu computadora.
                        </div>
                    </div>
                </div>
                """),
            unsafe_allow_html=True,
        )

        if st.button(
            "✓ Imagen seleccionada" if carga_activa else "Seleccionar imagen",
            key="seleccionar_carga",
            use_container_width=True,
            type="primary" if carga_activa else "secondary",
        ):
            st.session_state["metodo_inspeccion"] = "Cargar imagen"
            st.rerun()

    with col_metodo_2:
        camara_activa = st.session_state["metodo_inspeccion"] == "Tomar fotografía"

        st.markdown(
            dedent(f"""
                <div class="inspection-card {'active' if camara_activa else ''}">
                    <div class="inspection-card-top"></div>
                    <div class="inspection-card-content">
                        <div class="inspection-card-icon"><i class="bi bi-camera"></i></div>
                        <div class="inspection-card-title">Tomar fotografía</div>
                        <div class="inspection-card-description">
                            Captura una imagen utilizando la cámara.
                        </div>
                    </div>
                </div>
                """),
            unsafe_allow_html=True,
        )

        if st.button(
            "✓ Cámara seleccionada" if camara_activa else "Abrir cámara",
            key="seleccionar_camara",
            use_container_width=True,
            type="primary" if camara_activa else "secondary",
        ):
            st.session_state["metodo_inspeccion"] = "Tomar fotografía"
            st.rerun()
    with col_metodo_3:
        streaming_activo = (
            st.session_state["metodo_inspeccion"] == "Streaming simulado"
        )

        clase_activa = "active" if streaming_activo else ""

        html_streaming = (
            f'<div class="inspection-card {clase_activa}">'
            '<div class="inspection-card-top"></div>'
            '<div class="inspection-card-content">'
            '<div class="inspection-card-icon"><i class="bi bi-film"></i></div>'
            '<div class="inspection-card-title">Streaming simulado</div>'
            '<div class="inspection-card-description">'
            'Procesa varias imágenes de forma secuencial simulando una línea de inspección.'
            '</div>'
            '</div>'
            '</div>'
        )

        st.markdown(
            html_streaming,
            unsafe_allow_html=True
        )

        if st.button(
            "✓ Streaming seleccionado" if streaming_activo else "Iniciar streaming",
            key="seleccionar_streaming",
            use_container_width=True,
            type="primary" if streaming_activo else "secondary",
        ):
            st.session_state["metodo_inspeccion"] = "Streaming simulado"
            st.rerun()
    opcion = st.session_state["metodo_inspeccion"]

    if opcion == "Cargar imagen":
        st.success("📂 Método seleccionado: Cargar imagen")

    elif opcion == "Tomar fotografía":
        st.success("📷 Método seleccionado: Tomar fotografía")

    elif opcion == "Streaming simulado":
        st.success("🎞️ Método seleccionado: Streaming simulado")

    else:
        st.info("👆 Selecciona un método de inspección para comenzar.")
    # -------- PROCESAR Y REGISTRAR INSPECCIÓN --------

    def procesar_inspeccion(imagen, nombre_archivo, origen):

        os.makedirs("inspecciones", exist_ok=True)

        ruta_imagen = os.path.join("inspecciones", nombre_archivo)
        with open(ruta_imagen, "wb") as archivo:

            archivo.write(imagen.getbuffer())

        with st.spinner("La inteligencia artificial está analizando la pieza..."):

            respuesta_modelo = clasificar_imagen(ruta_imagen)

        # -------- INTERPRETAR RESPUESTA DEL MODELO --------

        if isinstance(respuesta_modelo, dict):

            resultado = respuesta_modelo.get("estado", "DESCONOCIDO")

            confianza = respuesta_modelo.get("confianza", 0.0)

            defecto = respuesta_modelo.get("defecto", None)

            resultado_yolo = respuesta_modelo.get("resultado_yolo")

            if resultado_yolo is not None:
                imagen_resultado = resultado_yolo.plot()
            else:
                imagen_resultado = None

        else:

            st.error("No fue posible interpretar la respuesta del modelo.")
            return

        # -------- MOSTRAR RESULTADO --------

        mostrar_encabezado_seccion(
            "Resultado de la inspección",
            "Clasificación y nivel de confianza obtenido por el modelo.",
        )

        resultado_normalizado = str(resultado).strip().upper()

        if resultado_normalizado in ["APTO", "BUENA"]:

            resultado_registro = "APTO"
            st.markdown("""
                <div style="
                    display:flex;
                    align-items:center;
                    gap:10px;
                    background:rgba(34, 197, 94, 0.12);
                    border:1px solid rgba(34, 197, 94, 0.35);
                    padding:12px 14px;
                    border-radius:7px;
                ">
                    <svg width="24" height="24" viewBox="0 0 24 24"
                        fill="none" stroke="#22c55e" stroke-width="2"
                        stroke-linecap="round" stroke-linejoin="round">
                        <circle cx="12" cy="12" r="9"></circle>
                        <path d="M8 12l2.5 2.5L16 9"></path>
                    </svg>
                    <span>
                        <strong>PIEZA APTA</strong> — La pieza cumple con los criterios de calidad.
                    </span>
                </div>
                """, unsafe_allow_html=True)

        elif resultado_normalizado in ["NO APTO", "MALA"]:

            resultado_registro = "NO APTO"
            st.markdown("""
                <div style="
                    display:flex;
                    align-items:center;
                    gap:10px;
                    background:rgba(239, 68, 68, 0.12);
                    border:1px solid rgba(239, 68, 68, 0.35);
                    padding:12px 14px;
                    border-radius:7px;
                ">
                    <svg width="24" height="24" viewBox="0 0 24 24"
                        fill="none" stroke="#ef4444" stroke-width="2"
                        stroke-linecap="round" stroke-linejoin="round">
                        <circle cx="12" cy="12" r="9"></circle>
                        <path d="M9 9l6 6M15 9l-6 6"></path>
                    </svg>
                    <span>
                        <strong>PIEZA NO APTA</strong> — La pieza requiere revisión o rechazo.
                    </span>
                </div>
                """, unsafe_allow_html=True)

            if defecto:
                st.markdown(
                    f"**Defecto detectado:** "
                    f"{str(defecto).replace('_', ' ').title()}"
                )

        else:

            resultado_registro = resultado_normalizado
            st.warning(f"⚠ Resultado: {resultado_normalizado}")

        try:
            confianza_numero = float(confianza)
        except (TypeError, ValueError):
            confianza_numero = 0.0

        # Si la confianza viene entre 0 y 1,
        # se convierte a porcentaje
        if confianza_numero <= 1:
            confianza_porcentaje = confianza_numero * 100
        else:
            confianza_porcentaje = confianza_numero

        # -------- GUARDAR EN DJANGO --------

        guardado_api, mensaje_api = guardar_inspeccion_api(
            resultado=resultado_registro,
            defecto=defecto or "",
            confianza=confianza_porcentaje,
            archivo=ruta_imagen,
            origen=origen,
        )

        if guardado_api:
            st.caption(mensaje_api)
        else:
            st.warning(mensaje_api)

            # -------- PANEL DE RESULTADOS --------

        col_imagen, col_info = st.columns([2, 1])

        with col_imagen:
            st.markdown("""
            <div style="display:flex; align-items:center; gap:9px; margin-bottom:12px;">
                <svg width="24" height="24" viewBox="0 0 24 24"
                    fill="none" stroke="#38bdf8" stroke-width="2"
                    stroke-linecap="round" stroke-linejoin="round">
                    <rect x="3" y="3" width="18" height="18" rx="2"></rect>
                    <circle cx="8.5" cy="8.5" r="1.5"></circle>
                    <path d="M21 15l-5-5L5 21"></path>
                </svg>
                <h3 style="margin:0;">Imagen procesada</h3>
            </div>
            """, unsafe_allow_html=True)

            if imagen_resultado is not None:
                try:
                    imagen_rgb = cv2.cvtColor(imagen_resultado, cv2.COLOR_BGR2RGB)

                    st.image(imagen_rgb, width=600)

                except Exception:
                    st.image(imagen_resultado, width=600)

        with col_info:

            with st.container(border=True):

                st.markdown("""
                <div style="display:flex; align-items:center; gap:9px; margin-bottom:10px;">
                    <svg width="24" height="24" viewBox="0 0 24 24"
                        fill="none" stroke="#38bdf8" stroke-width="2"
                        stroke-linecap="round" stroke-linejoin="round">
                        <path d="M4 19V10"></path>
                        <path d="M10 19V5"></path>
                        <path d="M16 19v-7"></path>
                        <path d="M22 19H2"></path>
                    </svg>
                    <h3 style="margin:0;">Confianza del modelo</h3>
                </div>
                """, unsafe_allow_html=True)

                st.metric(label="Resultado", value=f"{confianza_porcentaje:.2f}%")

                st.progress(min(max(confianza_porcentaje / 100, 0.0), 1.0))

            with st.container(border=True):

                st.markdown("""
                <div style="display:flex; align-items:center; gap:9px; margin-bottom:10px;">
                    <svg width="24" height="24" viewBox="0 0 24 24"
                        fill="none" stroke="#38bdf8" stroke-width="2"
                        stroke-linecap="round" stroke-linejoin="round">
                        <path d="M3 6h6l2 2h10v10a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V6z"></path>
                    </svg>
                    <h3 style="margin:0;">Origen de la imagen</h3>
                </div>
                """, unsafe_allow_html=True)

                st.metric(label="Fuente", value=origen)

            if defecto:

                with st.container(border=True):

                    defecto_texto = str(defecto).replace("_", " ").title()

                    st.markdown("""
                    <div style="display:flex; align-items:center; gap:9px; margin-bottom:10px;">
                        <svg width="24" height="24" viewBox="0 0 24 24"
                            fill="none" stroke="#f59e0b" stroke-width="2"
                            stroke-linecap="round" stroke-linejoin="round">
                            <path d="M10.3 3.6L2.5 17a2 2 0 0 0 1.7 3h15.6a2 2 0 0 0 1.7-3L13.7 3.6a2 2 0 0 0-3.4 0z"></path>
                            <path d="M12 9v4"></path>
                            <path d="M12 17h.01"></path>
                        </svg>
                        <h3 style="margin:0;">Defecto detectado</h3>
                    </div>
                    """, unsafe_allow_html=True)

                    st.metric(label="Clasificación", value=defecto_texto)

                    # -------- GUARDAR EN CSV --------

            fecha_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            nueva_fila = [
                fecha_hora,
                resultado_registro,
                defecto or "",
                f"{confianza_porcentaje:.2f}",
                nombre_archivo,
                origen,
            ]

            with open(archivo_csv, mode="a", newline="", encoding="utf-8") as archivo:

                escritor = csv.writer(archivo)
                escritor.writerow(nueva_fila)
        # -------- CARGAR IMAGEN --------

    if opcion == "Cargar imagen":

        archivo_subido = st.file_uploader(
            "Selecciona una imagen de la pieza",
            type=["jpg", "jpeg", "png"],
            key="imagen_cargada",
        )

        if archivo_subido is not None:

            analizar = st.button(
                "🔍 Analizar imagen",
                key="boton_analizar_archivo",
                use_container_width=True,
            )

            if analizar:

                nombre_archivo = datetime.now().strftime("archivo_%Y%m%d_%H%M%S.jpg")

                procesar_inspeccion(archivo_subido, nombre_archivo, "Archivo local")

            else:

                st.image(
                    archivo_subido,
                    caption="Imagen seleccionada",
                    use_container_width=True,
                )
    # -------- TOMAR FOTOGRAFÍA --------

    elif opcion == "Tomar fotografía":

        fotografia = st.camera_input(
            "Coloca la pieza frente a la cámara", key="fotografia_camara"
        )

        if fotografia is not None:

            if st.button(
                "📷 Analizar fotografía",
                key="boton_analizar_camara",
                use_container_width=True,
            ):

                nombre_archivo = datetime.now().strftime("inspeccion_%Y%m%d_%H%M%S.jpg")

                procesar_inspeccion(fotografia, nombre_archivo, "Cámara")
    # -------- STREAMING SIMULADO --------

    elif opcion == "Streaming simulado":

        imagenes_streaming = st.file_uploader(
            "Selecciona varias imágenes para simular el flujo de inspección",
            type=["jpg", "jpeg", "png"],
            accept_multiple_files=True,
            key="streaming_procesamiento",
        )

        if imagenes_streaming:

            st.info(
                f"🎞️ {len(imagenes_streaming)} imágenes listas para la simulación."
            )

            iniciar_simulacion = st.button(
                "▶ Ejecutar simulación",
                key="boton_ejecutar_streaming",
                use_container_width=True,
            )

            if iniciar_simulacion:

                barra_progreso = st.progress(0)

                total_imagenes = len(imagenes_streaming)

                for indice, imagen_stream in enumerate(
                    imagenes_streaming,
                    start=1
                ):

                    st.markdown(
                        f"### 🎞️ Inspección {indice} de {total_imagenes}"
                    )

                    nombre_archivo = datetime.now().strftime(
                        f"stream_{indice}_%Y%m%d_%H%M%S_%f.jpg"
                    )

                    procesar_inspeccion(
                        imagen_stream,
                        nombre_archivo,
                        "Streaming simulado",
                    )

                    barra_progreso.progress(
                        indice / total_imagenes
                    )

                st.success(
                    f"✅ Simulación finalizada. "
                    f"{total_imagenes} imágenes procesadas."
                )
    st.divider()


def mostrar_resumen(total, aptas, no_aptas):

    mostrar_encabezado_seccion(
        "Resumen general", "Indicadores principales de las inspecciones registradas."
    )

    porcentaje_aptas = (aptas / total) * 100 if total > 0 else 0

    porcentaje_no_aptas = (no_aptas / total) * 100 if total > 0 else 0

    col1, col2, col3 = st.columns(3)

    tarjetas = [
        {
            "icono": "bi bi-clipboard-data",
            "valor": total,
            "titulo": "Total de inspecciones",
            "descripcion": "Registros almacenados",
            "detalle": "Base de datos actualizada",
            "clase": "kpi-total",
        },
        {
            "icono": "bi bi-check-circle",
            "valor": aptas,
            "titulo": "Piezas aptas",
            "descripcion": "Cumplen con calidad",
            "detalle": f"{porcentaje_aptas:.1f}% del total",
            "clase": "kpi-success",
        },
        {
            "icono": "bi bi-exclamation-triangle",
            "valor": no_aptas,
            "titulo": "Piezas no aptas",
            "descripcion": "Requieren revisión",
            "detalle": f"{porcentaje_no_aptas:.1f}% del total",
            "clase": "kpi-danger",
        },
    ]

    for columna, tarjeta in zip([col1, col2, col3], tarjetas):

        with columna:

            html = (
                f'<div class="kpi-card {tarjeta["clase"]}">'
                '<div class="kpi-card-top">'
                '<div class="kpi-icon">'
                f'<i class="{tarjeta["icono"]}"></i>'
                "</div>"
                '<div class="kpi-detail">'
                f'{tarjeta["detalle"]}'
                "</div>"
                "</div>"
                '<div class="kpi-value">'
                f'{tarjeta["valor"]}'
                "</div>"
                '<div class="kpi-title">'
                f'{tarjeta["titulo"]}'
                "</div>"
                '<div class="kpi-description">'
                f'{tarjeta["descripcion"]}'
                "</div>"
                "</div>"
            )

            st.markdown(html, unsafe_allow_html=True)

    st.markdown(
        '<div style="font-size:21px; color:#6b7280; margin-top:4px;">Actualización automática basada en las inspecciones registradas.</div>',
        unsafe_allow_html=True
    )
    st.divider()


def mostrar_graficas(aptas, no_aptas):
    tema_oscuro = st.session_state.get("tema", "Claro") == "Oscuro"

    fondo_grafica = "#172033" if tema_oscuro else "#FFFFFF"
    color_texto = "#F8FAFC" if tema_oscuro else "#231F20"
    color_ejes = "#AAB6C8" if tema_oscuro else "#555555"
    mostrar_encabezado_seccion(
        "Análisis de inspección", "Comparación visual entre piezas aptas y no aptas."
    )

    col1, col2 = st.columns(2)

    etiquetas = ["Aptas", "No aptas"]
    valores = [aptas, no_aptas]

    with col1:

        st.markdown("### Resultados de inspección")

        fig_barras, ax_barras = plt.subplots(facecolor=fondo_grafica)
        ax_barras.set_facecolor(fondo_grafica)

        ax_barras.tick_params(colors=color_ejes)
        ax_barras.xaxis.label.set_color(color_texto)
        ax_barras.yaxis.label.set_color(color_texto)
        ax_barras.title.set_color(color_texto)

        for spine in ax_barras.spines.values():
            spine.set_color(color_ejes)

        barras = ax_barras.bar(
            etiquetas,
            valores,
            color=["#1D7EAE", "#FF661B"]
        )

        ax_barras.set_ylabel("Cantidad")
        ax_barras.set_title("Piezas inspeccionadas")

        for barra, valor in zip(barras, valores):

            ax_barras.text(
                barra.get_x() + barra.get_width() / 2,
                barra.get_height(),
                str(valor),
                ha="center",
                va="bottom",
                color=color_texto,
            )

        st.pyplot(fig_barras)

    with col2:

        st.markdown("### Distribución de resultados")

        total = aptas + no_aptas

        if total > 0:

            fig_dona, ax_dona = plt.subplots(facecolor=fondo_grafica)
            ax_dona.set_facecolor(fondo_grafica)

            ax_dona.pie(
                valores,
                labels=etiquetas,
                colors=["#1D7EAE", "#FF661B"],
                autopct="%1.1f%%",
                startangle=90,
                wedgeprops={"width": 0.40},
                textprops={"color": color_texto},
            )

            ax_dona.axis("equal")

            st.pyplot(fig_dona)

        else:

            st.info("Todavía no hay inspecciones para mostrar la distribución.")

    st.divider()


def mostrar_indicadores(aptas, no_aptas):

    inspecciones_validas = aptas + no_aptas

    if inspecciones_validas == 0:
        st.info("Todavía no hay inspecciones suficientes para calcular indicadores.")
        st.divider()
        return

    porcentaje_apto = (aptas / inspecciones_validas) * 100

    porcentaje_no_apto = (no_aptas / inspecciones_validas) * 100

    mostrar_encabezado_seccion(
        "Indicadores de calidad", "Métricas porcentuales del desempeño del proceso."
    )

    col4, col5 = st.columns(2)

    with col4:
        st.markdown(
            f"""
            <div class="kpi-card">
                <div class="kpi-icon"><i class="bi bi-graph-up-arrow"></i></div>
                <div class="kpi-value">{porcentaje_apto:.1f}%</div>
                <div class="kpi-label">Tasa de aprobación</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col5:
        st.markdown(
            f"""
            <div class="kpi-card">
                <div class="kpi-icon"><i class="bi bi-graph-down-arrow"></i></div>
                <div class="kpi-value">{porcentaje_no_apto:.1f}%</div>
                <div class="kpi-label">Tasa de rechazo</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.divider()


def mostrar_registro():

    url_api = "http://127.0.0.1:8000/api/inspecciones/"

    try:
        respuesta = requests.get(url_api, timeout=10)

        if respuesta.status_code != 200:
            st.error("No fue posible consultar el registro en Django.")
            st.caption(f"Error de API: {respuesta.status_code}")
            st.divider()
            return

        registros = respuesta.json()

        if not registros:
            st.info("Todavía no hay inspecciones registradas.")
            st.divider()
            return

        datos = pd.DataFrame(registros)

        # Cambiar nombres de columnas
        datos = datos.rename(
            columns={
                "fecha": "Fecha",
                "resultado": "Resultado",
                "defecto": "Defecto",
                "confianza": "Confianza (%)",
                "archivo": "Archivo Guardado",
                "origen": "Origen",
            }
        )

        # Convertir y ordenar fechas
        datos["Fecha"] = pd.to_datetime(
            datos["Fecha"],
            errors="coerce",
        )

        datos = datos.sort_values(
            by="Fecha",
            ascending=True,
        ).reset_index(drop=True)

        datos["Fecha"] = datos["Fecha"].dt.strftime("%d/%m/%Y %H:%M:%S")

        # -------- ÚLTIMA INSPECCIÓN --------

        ultima_inspeccion = datos.iloc[-1]

        st.markdown("### Última inspección")

        resultado = str(ultima_inspeccion["Resultado"]).strip().upper()

        defecto = ultima_inspeccion.get("Defecto", "")

        if pd.notna(defecto) and str(defecto).strip():
            defecto_texto = str(defecto).replace("_", " ").title()
        else:
            defecto_texto = "Sin defecto"

        try:
            confianza = float(ultima_inspeccion["Confianza (%)"])
        except (TypeError, ValueError):
            confianza = 0.0

        col_resultado, col_defecto, col_confianza = st.columns(3)

        # -------- PREPARAR ESTADO VISUAL --------

        if resultado in ["APTO", "BUENA"]:
            clase_resultado = "registro-apto"
            icono_resultado = "✅"
            texto_resultado = "PIEZA APTA"

        elif resultado in ["NO APTO", "MALA"]:
            clase_resultado = "registro-no-apto"
            icono_resultado = "❌"
            texto_resultado = "PIEZA NO APTA"

        else:
            clase_resultado = "registro-revision"
            icono_resultado = "⚠️"
            texto_resultado = resultado

        # -------- TARJETA RESULTADO --------

        with col_resultado:
            st.markdown(
                f"""<div class="registro-card {clase_resultado}">
<div class="registro-card-top"></div>
<div class="registro-card-content">
<div class="registro-card-heading"><i class="bi bi-clipboard-check"></i> Resultado</div>
<div class="registro-result-badge">{icono_resultado} {texto_resultado}</div>
<div class="registro-card-caption">Clasificación general de la última inspección.</div>
<div class="registro-card-caption">
    <i class="bi bi-calendar3"></i>
    <strong>Fecha:</strong> {ultima_inspeccion['Fecha']}
</div>
</div>
</div>""",
                unsafe_allow_html=True,
            )

        # -------- TARJETA DEFECTO --------

        with col_defecto:
            st.markdown(
                f"""<div class="registro-card">
<div class="registro-card-top"></div>
<div class="registro-card-content">
<div class="registro-card-heading"><i class="bi bi-exclamation-triangle"></i> Defecto</div>
<div class="registro-card-label">Clasificación detectada</div>
<div class="registro-card-value">{escape(defecto_texto)}</div>
<div class="registro-card-caption">Tipo de defecto identificado por el modelo.</div>
<div class="registro-card-caption">
    <i class="bi bi-folder2-open"></i>
    <strong>Origen:</strong> {ultima_inspeccion['Origen']}
</div>
</div>
</div>""",
                unsafe_allow_html=True,
            )

        # -------- TARJETA CONFIANZA --------

        with col_confianza:
            porcentaje_barra = min(
                max(confianza, 0.0),
                100.0,
            )

            st.markdown(
                f"""<div class="registro-card">
<div class="registro-card-top"></div>
<div class="registro-card-content">
<div class="registro-card-heading"><i class="bi bi-shield-check"></i> Confianza</div>
<div class="registro-card-label">Nivel del modelo</div>
<div class="registro-card-value">{confianza:.2f}%</div>
<div class="registro-progress-track">
<div class="registro-progress-fill" style="width:{porcentaje_barra:.2f}%;"></div>
</div>
</div>
</div>""",
                unsafe_allow_html=True,
            )

        # -------- HISTORIAL COMPLETO --------

        with st.container(border=True):
            st.markdown(
                """
                <h3 class="registro-section-title">
                    <i class="bi bi-clipboard-data"></i> Historial de inspecciones
                </h3>
                """,
                unsafe_allow_html=True,
            )

            columnas_mostradas = [
                "Fecha",
                "Resultado",
                "Defecto",
                "Confianza (%)",
                "Origen",
                "Archivo Guardado",
            ]

            datos_mostrados = (
                datos[columnas_mostradas].iloc[::-1].reset_index(drop=True)
            )

            # Guardar resultado original para futuros colores
            resultados_originales = (
                datos_mostrados["Resultado"].astype(str).str.strip().str.upper()
            )

            # Mejorar visualización del resultado
            datos_mostrados["Resultado"] = resultados_originales.replace(
                {
                    "APTO": "🟢 APTO",
                    "BUENA": "🟢 APTO",
                    "NO APTO": "🔴 NO APTO",
                    "MALA": "🔴 NO APTO",
                }
            )

            # Mejorar visualización del defecto
            datos_mostrados["Defecto"] = (
                datos_mostrados["Defecto"]
                .fillna("")
                .astype(str)
                .str.strip()
            )

            datos_mostrados.loc[
                datos_mostrados["Defecto"].isin(["", "nan", "None"]),
                "Defecto"
            ] = "Sin defecto"

            # Formato de confianza
            datos_mostrados["Confianza (%)"] = (
                pd.to_numeric(
                    datos_mostrados["Confianza (%)"],
                    errors="coerce",
                )
                .fillna(0)
                .round(2)
            )
            # Color de barra según resultado
            datos_mostrados["_Color"] = resultados_originales.apply(
                lambda x: "#22c55e" if x in ["APTO", "BUENA"] else "#ef4444"
            )

            filas_html = ""

            for _, fila in datos_mostrados.iterrows():
                confianza_fila = float(fila["Confianza (%)"])
                color_barra = fila["_Color"]
                # Crear miniatura de la imagen de la inspección
                nombre_imagen = str(fila["Archivo Guardado"]).strip()

                ruta_imagen = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "inspecciones",
                    os.path.basename(nombre_imagen)
                )

                imagen_html = "<span>Sin imagen</span>"

                if os.path.exists(ruta_imagen):
                    try:
                        with Image.open(ruta_imagen) as img:
                            img.thumbnail((80, 80))

                            if img.mode not in ("RGB", "L"):
                                img = img.convert("RGB")

                            buffer = io.BytesIO()
                            img.save(buffer, format="JPEG", quality=70)

                            imagen_base64 = base64.b64encode(
                                buffer.getvalue()
                            ).decode("utf-8")

                        imagen_html = f"""
                        <img
                            src="data:image/jpeg;base64,{imagen_base64}"
                            style="
                                width:60px;
                                height:60px;
                                object-fit:cover;
                                border-radius:8px;
                                border:1px solid #d1d5db;
                            "
                            alt="Miniatura de inspección"
                        >
                        """

                    except Exception:
                        imagen_html = "<span>No disponible</span>"

                else:
                    imagen_html = "<span>No disponible</span>"
                filas_html += dedent(f"""
                <tr>
                    <td>{fila["Fecha"]}</td>
                    <td>{fila["Resultado"]}</td>
                    <td>{fila["Defecto"]}</td>
                    <td>{imagen_html}</td>
                    <td>
                        <div style="
                            display:flex;
                            align-items:center;
                            gap:12px;
                        ">
                            <div style="
                                flex:1;
                                height:8px;
                                background:#374151;
                                border-radius:20px;
                                overflow:hidden;
                            ">
                                <div style="
                                    width:{confianza_fila}%;
                                    height:100%;
                                    background:{color_barra};
                                    border-radius:20px;
                                "></div>
                            </div>

                            <span style="
                                min-width:58px;
                                text-align:right;
                                font-weight:600;
                            ">
                                {confianza_fila:.2f}%
                            </span>
                        </div>
                    </td>
                    <td>{fila["Origen"]}</td>
                </tr>
                """)
            tema_actual = st.session_state.get("tema", "Claro")

            if tema_actual == "Oscuro":
                fondo_tabla = "#111827"
                fondo_encabezado = "#171c26"
                color_texto = "#f8fafc"
                borde_tabla = "#374151"
                borde_filas = "#2b3443"
                fondo_hover = "#182235"
            else:
                fondo_tabla = "#ffffff"
                fondo_encabezado = "#f8fafc"
                color_texto = "#1f2937"
                borde_tabla = "#dbe5ee"
                borde_filas = "#e5e7eb"
                fondo_hover = "#f3f6f9"
            tabla_html = dedent(f"""
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <style>
                body {{
                    margin: 0;
                    padding: 0;
                    background: transparent;
                    font-family: Arial, sans-serif;
                }}

                .historial-box {{
                    background: {fondo_tabla};
                    border: 1px solid {borde_tabla};
                    border-radius: 12px;
                    overflow: hidden;
                    color: {color_texto};
                }}

                .historial-scroll {{
                    max-height: 510px;
                    overflow-y: auto;
                }}

                table {{
                    width: 100%;
                    border-collapse: collapse;
                    font-size: 16px;
                    color: {color_texto};
                }}

                thead {{
                    position: sticky;
                    top: 0;
                    background: #0879C9;
                    color: white;
                    z-index: 2;
                }}

                thead th {{
                    background: #0879C9;
                    color: white;
                    font-size: 15px;
                    font-weight: 600;
                    padding: 14px 12px !important;
                    border-bottom: none;
                }}

                tbody td {{
                    font-size: 16px;
                    padding: 14px 12px;
                }}

                th {{
                    padding: 12px;
                    text-align: left;
                    border-bottom: 1px solid {borde_tabla};
                    color: {color_texto};
                }}

                td {{
                    padding: 10px 12px;
                    border-bottom: 1px solid {borde_filas};
                    color: {color_texto};
                }}

                tbody tr {{
                    background: {fondo_tabla};
                }}

                tbody tr:hover {{
                    background: {fondo_hover};
                }}
                
            </style>

            <div class="historial-box">
                <div class="historial-scroll">
                    <table>

                <thead style="
                    position:sticky;
                    top:0;
                    background:{fondo_encabezado};
                    color:{color_texto};
                    z-index:2;
                ">
                    <tr>
                        <th style="padding:12px;text-align:left;"><i class="bi bi-calendar3"></i> Fecha</th>
                        <th style="padding:12px;text-align:left;"><i class="bi bi-check2-circle"></i> Resultado</th>
                        <th style="padding:12px;text-align:left;"><i class="bi bi-exclamation-triangle"></i> Defecto</th>
                        <th style="padding:12px;text-align:left;"><i class="bi bi-image"></i> Imagen</th>
                        <th style="padding:12px;text-align:left;"><i class="bi bi-shield-check"></i> Confianza</th>
                        <th style="padding:12px;text-align:left;"><i class="bi bi-folder2-open"></i> Origen</th>
                    </tr>
                </thead>

                <tbody>
                    {filas_html}
                </tbody>
                        </table>
                    </div>
                </div>
                """)

            components.html(
                tabla_html,
                height=520,
                scrolling=True,
            )
            st.info(f"📊 Total de inspecciones registradas: **{len(datos)}**")

    except requests.exceptions.ConnectionError:
        st.error("No se pudo conectar con Django.")
        st.info("Verifica que esté ejecutándose: " "`python manage.py runserver`")

    except requests.exceptions.RequestException as error:
        st.error("Ocurrió un error al consultar la API.")
        st.caption(f"Detalle técnico: {error}")

    except Exception as error:
        st.error("No fue posible cargar el registro de inspecciones.")
        st.caption(f"Detalle técnico: {error}")

    st.divider()


def generar_excel_registros(registros):

    datos = pd.DataFrame(registros)

    datos = datos.rename(
        columns={
            "fecha": "Fecha",
            "resultado": "Resultado",
            "defecto": "Defecto",
            "confianza": "Confianza (%)",
            "archivo": "Archivo Guardado",
            "origen": "Origen",
        }
    )

    columnas = [
        "Fecha",
        "Resultado",
        "Defecto",
        "Confianza (%)",
        "Archivo Guardado",
        "Origen",
    ]

    datos = datos[columnas]

    wb = Workbook()
    ws = wb.active
    ws.title = "Registro VisionQA"

    # -------- TÍTULO --------

    ws.merge_cells("A1:F1")
    ws["A1"] = "VisionQA - Historial de Inspecciones"

    ws["A1"].font = Font(
        size=16,
        bold=True,
        color="FFFFFF",
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws["A1"].fill = PatternFill(
        fill_type="solid",
        start_color="1F77B4",
        end_color="1F77B4",
    )

    # -------- SUBTÍTULO --------

    ws.merge_cells("A2:F2")

    ws["A2"] = (
        "Sistema Inteligente de Inspección Visual "
        "Asistido por Inteligencia Artificial"
    )

    ws["A2"].font = Font(
        italic=True,
        size=10,
    )

    ws["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    # -------- INFORMACIÓN --------

    ws["A4"] = "Fecha de generación:"
    ws["B4"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws["D4"] = "Registros:"
    ws["E4"] = len(datos)

    borde_delgado = Border(
        left=Side(style="thin", color="D9E1E8"),
        right=Side(style="thin", color="D9E1E8"),
        top=Side(style="thin", color="D9E1E8"),
        bottom=Side(style="thin", color="D9E1E8"),
    )

    # -------- ENCABEZADOS --------

    fila_encabezado = 6

    for numero_columna, nombre in enumerate(
        columnas,
        start=1,
    ):

        celda = ws.cell(
            row=fila_encabezado,
            column=numero_columna,
            value=nombre,
        )

        celda.font = Font(
            bold=True,
            color="FFFFFF",
        )

        celda.fill = PatternFill(
            fill_type="solid",
            start_color="1F7FB4",
            end_color="1F7FB4",
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        celda.border = borde_delgado

    # -------- REGISTROS --------

    fila_actual = fila_encabezado + 1

    for fila in datos.itertuples(index=False):

        for numero_columna, valor in enumerate(
            fila,
            start=1,
        ):

            celda = ws.cell(
                row=fila_actual,
                column=numero_columna,
                value=valor,
            )

            celda.border = borde_delgado

            celda.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

            if numero_columna == 2:

                resultado = str(valor).strip().upper()

                if resultado == "APTO":

                    celda.fill = PatternFill(
                        fill_type="solid",
                        start_color="D9EAD3",
                        end_color="D9EAD3",
                    )

                    celda.font = Font(
                        bold=True,
                        color="008000",
                    )

                elif resultado == "NO APTO":

                    celda.fill = PatternFill(
                        fill_type="solid",
                        start_color="F4CCCC",
                        end_color="F4CCCC",
                    )

                    celda.font = Font(
                        bold=True,
                        color="C00000",
                    )

        fila_actual += 1

    # -------- AJUSTAR COLUMNAS --------

    from openpyxl.utils import get_column_letter

    for indice_columna in range(1, 7):

        letra = get_column_letter(indice_columna)

        longitud = 0

        for fila in ws.iter_rows(
            min_row=6,
            max_row=fila_actual - 1,
            min_col=indice_columna,
            max_col=indice_columna,
        ):

            celda = fila[0]

            if celda.value is not None:

                longitud = max(longitud, len(str(celda.value)))

        ws.column_dimensions[letra].width = min(
            longitud + 4,
            35,
        )

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:F{fila_actual - 1}"

    archivo_excel = BytesIO()

    wb.save(archivo_excel)
    archivo_excel.seek(0)

    return archivo_excel.getvalue(), len(datos)


def limpiar_texto_pdf(texto):

    texto = str(texto)

    # Eliminar símbolos Markdown.
    texto = texto.replace("**", "")
    texto = texto.replace("__", "")
    texto = texto.replace("`", "")

    # Eliminar emojis y símbolos que Helvetica no puede mostrar.
    texto = re.sub(
        r"[\U00010000-\U0010ffff]",
        "",
        texto,
    )

    return texto.strip()


def generar_pdf_informe_ia(
    resultado_ia,
    estado,
    confianza,
    prioridad,
    defecto,
    accion,
    tipo_analisis,
):

    buffer = io.BytesIO()

    documento = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=1.7 * cm,
        leftMargin=1.7 * cm,
        topMargin=1.7 * cm,
        bottomMargin=1.7 * cm,
        title="Informe Ejecutivo VisionQA",
        author="VisionQA",
    )

    estilos = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        name="TituloVisionQA",
        parent=estilos["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0F6CBD"),
        spaceAfter=8,
    )

    estilo_subtitulo = ParagraphStyle(
        name="SubtituloVisionQA",
        parent=estilos["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=13,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#667085"),
        spaceAfter=16,
    )

    estilo_seccion = ParagraphStyle(
        name="SeccionVisionQA",
        parent=estilos["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#172033"),
        spaceBefore=12,
        spaceAfter=7,
    )

    estilo_subseccion = ParagraphStyle(
        name="SubseccionVisionQA",
        parent=estilos["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=15,
        textColor=colors.HexColor("#29384D"),
        spaceBefore=8,
        spaceAfter=4,
    )

    estilo_texto = ParagraphStyle(
        name="TextoVisionQA",
        parent=estilos["BodyText"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=14,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#303846"),
        spaceAfter=5,
    )

    estilo_vineta = ParagraphStyle(
        name="VinetaVisionQA",
        parent=estilo_texto,
        leftIndent=12,
        firstLineIndent=-7,
        bulletIndent=3,
        spaceAfter=3,
    )

    elementos = []

    # ---------------------------------------------------------
    # ENCABEZADO
    # ---------------------------------------------------------

    elementos.append(
        Paragraph(
            "VisionQA - Informe Ejecutivo IA",
            estilo_titulo,
        )
    )

    elementos.append(
        Paragraph(
            "Sistema Inteligente de Inspeccion Visual Asistido "
            "por Inteligencia Artificial",
            estilo_subtitulo,
        )
    )

    informacion = [
        [
            Paragraph("<b>Tipo de analisis</b>", estilo_texto),
            Paragraph(
                escape(limpiar_texto_pdf(tipo_analisis)),
                estilo_texto,
            ),
        ],
        [
            Paragraph("<b>Fecha de generacion</b>", estilo_texto),
            Paragraph(
                datetime.now().strftime("%d/%m/%Y %H:%M"),
                estilo_texto,
            ),
        ],
    ]

    tabla_informacion = Table(
        informacion,
        colWidths=[4.2 * cm, 12.5 * cm],
    )

    tabla_informacion.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (0, -1),
                    colors.HexColor("#EAF3FA"),
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.6,
                    colors.HexColor("#C9D5E2"),
                ),
                (
                    "INNERGRID",
                    (0, 0),
                    (-1, -1),
                    0.3,
                    colors.HexColor("#D9E1E8"),
                ),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    elementos.append(tabla_informacion)
    elementos.append(Spacer(1, 14))

    # ---------------------------------------------------------
    # INDICADORES
    # ---------------------------------------------------------

    if "10 inspecciones" in tipo_analisis.lower():

        elementos.append(
            Paragraph(
                "Resumen del historial",
                estilo_seccion,
            )
        )

        registros_api = obtener_inspecciones_api()
        ultimas_10 = registros_api[:10] if registros_api else []

        total_historial = len(ultimas_10)

        aptas_historial = sum(
            1 for r in ultimas_10
            if str(r.get("resultado", "")).strip().upper() == "APTO"
        )

        no_aptas_historial = sum(
            1 for r in ultimas_10
            if str(r.get("resultado", "")).strip().upper() == "NO APTO"
        )

        tasa_rechazo = (
            (no_aptas_historial / total_historial) * 100
            if total_historial > 0
            else 0
        )

        defectos_historial = []

        for r in ultimas_10:
            defecto_actual = str(r.get("defecto", "")).strip()

            if (
                defecto_actual
                and defecto_actual.lower() not in ["sin defecto", "none", "nan"]
            ):
                defecto_limpio = defecto_actual.replace("_", " ").title()

                if defecto_limpio not in defectos_historial:
                    defectos_historial.append(defecto_limpio)

        defectos_texto = (
            ", ".join(defectos_historial)
            if defectos_historial
            else "Sin defectos registrados"
        )

        datos_indicadores = [
            [
                Paragraph("<b>Total</b>", estilo_texto),
                Paragraph("<b>Aptas</b>", estilo_texto),
                Paragraph("<b>No aptas</b>", estilo_texto),
                Paragraph("<b>Tasa de rechazo</b>", estilo_texto),
                Paragraph("<b>Defectos detectados</b>", estilo_texto),
            ],
            [
                Paragraph(str(total_historial), estilo_texto),
                Paragraph(str(aptas_historial), estilo_texto),
                Paragraph(str(no_aptas_historial), estilo_texto),
                Paragraph(f"{tasa_rechazo:.1f}%", estilo_texto),
                Paragraph(
                    escape(limpiar_texto_pdf(defectos_texto)),
                    estilo_texto,
                ),
            ],
        ]

    else:

        elementos.append(
            Paragraph(
                "Resumen de la inspección",
                estilo_seccion,
            )
        )

        datos_indicadores = [
            [
                Paragraph("<b>Resultado</b>", estilo_texto),
                Paragraph("<b>Confianza</b>", estilo_texto),
                Paragraph("<b>Prioridad</b>", estilo_texto),
                Paragraph("<b>Defecto</b>", estilo_texto),
                Paragraph("<b>Acción</b>", estilo_texto),
            ],
            [
                Paragraph(
                    escape(limpiar_texto_pdf(estado)),
                    estilo_texto,
                ),
                Paragraph(
                    escape(f"{confianza:.1f}%"),
                    estilo_texto,
                ),
                Paragraph(
                    escape(limpiar_texto_pdf(prioridad)),
                    estilo_texto,
                ),
                Paragraph(
                    escape(limpiar_texto_pdf(defecto)),
                    estilo_texto,
                ),
                Paragraph(
                    escape(limpiar_texto_pdf(accion)),
                    estilo_texto,
                ),
            ],
        ]

    tabla_indicadores = Table(
        datos_indicadores,
        colWidths=[
            3.15 * cm,
            3.15 * cm,
            3.15 * cm,
            3.65 * cm,
            3.65 * cm,
        ],
    )

    tabla_indicadores.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#DCECF8"),
                ),
                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, 1),
                    colors.white,
                ),
                (
                    "BOX",
                    (0, 0),
                    (-1, -1),
                    0.7,
                    colors.HexColor("#BFCBD8"),
                ),
                (
                    "INNERGRID",
                    (0, 0),
                    (-1, -1),
                    0.4,
                    colors.HexColor("#D9E1E8"),
                ),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    elementos.append(tabla_indicadores)
    elementos.append(Spacer(1, 14))

    # ---------------------------------------------------------
    # CONTENIDO GENERADO POR GEMINI
    # ---------------------------------------------------------

    lineas = str(resultado_ia).splitlines()

    for linea in lineas:

        linea = limpiar_texto_pdf(linea)

        if not linea:
            continue

        if linea.startswith("# "):

            titulo = linea[2:].strip()

            elementos.append(
                Paragraph(
                    escape(titulo),
                    estilo_seccion,
                )
            )

        elif linea.startswith("## "):

            titulo = linea[3:].strip()

            elementos.append(
                Paragraph(
                    escape(titulo),
                    estilo_seccion,
                )
            )

        elif linea.startswith("### "):

            titulo = linea[4:].strip()

            elementos.append(
                Paragraph(
                    escape(titulo),
                    estilo_subseccion,
                )
            )

        elif linea.startswith("- "):

            contenido = linea[2:].strip()

            elementos.append(
                Paragraph(
                    f"- {escape(contenido)}",
                    estilo_vineta,
                )
            )

        elif re.match(r"^\d+\.\s", linea):

            elementos.append(
                Paragraph(
                    escape(linea),
                    estilo_vineta,
                )
            )

        elif linea == "---":

            elementos.append(Spacer(1, 5))

        else:

            elementos.append(
                Paragraph(
                    escape(linea),
                    estilo_texto,
                )
            )

    elementos.append(Spacer(1, 18))

    elementos.append(
        Paragraph(
            "Informe generado por VisionQA mediante Gemini IA. "
            "Proyecto desarrollado para IOT Technologies.",
            estilo_subtitulo,
        )
    )

    documento.build(elementos)

    buffer.seek(0)

    return buffer.getvalue()


def mostrar_gemini():
    st.markdown(
        '<span class="gemini-page-marker"></span>',
        unsafe_allow_html=True,
    )
    # -------- VARIABLES DEL INFORME --------

    resultado_ia = None
    titulo_informe = None
    descripcion_informe = None
    mostrar_tarjetas = False

    # -------- VARIABLES DE LAS TARJETAS --------

    estado_tarjeta = "Sin datos"
    confianza_tarjeta = 0.0
    prioridad_tarjeta = "Sin definir"
    defecto_tarjeta = "Sin defecto"
    accion_tarjeta = "Sin recomendación"

    # =========================================================
    # PANTALLA INICIAL
    # =========================================================

    if resultado_ia is None:

        # -------- FUNCIONES DEL MÓDULO --------
        st.markdown(
            """
            <style>
            .funcion-card {
                background: transparent;
                border: none;
                padding: 0;
                min-height: 105px;
            }

            .funcion-titulo {
                padding: 10px 14px;
                border-radius: 6px;
                font-size: 21px;
                font-weight: 600;
                margin-bottom: 14px;
            }

            .funcion-titulo i {
                margin-right: 7px;
                font-size: 21px;
            }

            .funcion-azul {
                background: #dbeafe;
                color: #1e3a5f;
            }

            .funcion-verde {
                background: #dcfce7;
                color: #166534;
            }

            .funcion-amarillo {
                background: #fef9c3;
                color: #854d0e;
            }

            .funcion-texto {
                font-size: 20px;
                color: #374151;
                line-height: 1.4;
            }
            div[data-testid="stButton"] button p {
                font-size: 20px !important;
                font-weight: 600 !important;
            }
            </style>
            """,
            unsafe_allow_html=True
        )
        st.markdown("### ¿Qué puede hacer este módulo?")

        funcion1, funcion2, funcion3, funcion4 = st.columns(4)

        with funcion1:
            st.markdown(
                """
                <div class="funcion-card">
                    <div class="funcion-titulo funcion-azul">
                        <i class="bi bi-search"></i>
                        Detectar patrones
                    </div>
                    <div class="funcion-texto">
                        Identifica defectos repetitivos y posibles tendencias.
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

        with funcion2:
            st.markdown(
                """
                <div class="funcion-card">
                    <div class="funcion-titulo funcion-verde">
                        <i class="bi bi-gear"></i>
                        Analizar causas
                    </div>
                    <div class="funcion-texto">
                        Organiza las causas potenciales mediante la metodología 6M.
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

        with funcion3:
            st.markdown(
                """
                <div class="funcion-card">
                    <div class="funcion-titulo funcion-amarillo">
                        <i class="bi bi-lightning-charge"></i>
                        Recomendar acciones
                    </div>
                    <div class="funcion-texto">
                        Propone acciones correctivas y oportunidades de mejora.
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

        with funcion4:
            st.markdown(
                """
                <div class="funcion-card">
                    <div class="funcion-titulo funcion-azul">
                        <i class="bi bi-file-earmark-text"></i>
                        Generar informes
                    </div>
                    <div class="funcion-texto">
                        Presenta los resultados en un formato ejecutivo y claro.
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )
        # -------- PROCESO DEL ANÁLISIS --------

        st.markdown("### Proceso del análisis")

        # Estilos de los pasos
        st.markdown(
            """
            <style>
            .paso-analisis {
                background: transparent;
                border: none;
                padding: 12px 10px;
                min-height: 90px;
            }

            .paso-titulo {
                font-size: 18px;
                font-weight: 600;
                color: #111827;
                margin-bottom: 14px;
                display: flex;
                align-items: center;
                gap: 8px;
            }

            .paso-numero {
                background: #3b82f6;
                color: white;
                width: 25px;
                height: 25px;
                border-radius: 5px;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                font-size: 14px;
                font-weight: 700;
            }

            .paso-texto {
                font-size: 20px;
                color: #6b7280;
                line-height: 1.4;
            }

            .flecha-proceso {
                height: 110px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 28px;
                color: #374151;
            }
            /* Texto dentro de mensajes success e info */
            div[data-testid="stAlert"] p {
                font-size: 21px !important;
            }
            div[data-testid="stCaptionContainer"] p {
                font-size: 19px !important;
            }
            </style>
            """,
            unsafe_allow_html=True
        )

        paso1, flecha1, paso2, flecha2, paso3, flecha3, paso4 = st.columns(
            [3, 0.7, 3, 0.7, 3, 0.7, 3]
        )
        with paso1:
            st.markdown(
                """
                <div class="paso-analisis">
                    <div class="paso-titulo">
                        <span class="paso-numero">1</span>
                        Selección
                    </div>
                    <div class="paso-texto">
                        El usuario selecciona el tipo de análisis.
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

        with flecha1:
            st.markdown(
                '<div class="flecha-proceso"><i class="bi bi-arrow-right"></i></div>',
                unsafe_allow_html=True
            )

        with paso2:
            st.markdown(
                """
                <div class="paso-analisis">
                    <div class="paso-titulo">
                        <span class="paso-numero">2</span>
                        Consulta
                    </div>
                    <div class="paso-texto">
                        VisionQA obtiene los registros almacenados.
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

        with flecha2:
            st.markdown(
                '<div class="flecha-proceso"><i class="bi bi-arrow-right"></i></div>',
                unsafe_allow_html=True
            )

        with paso3:
            st.markdown(
                """
                <div class="paso-analisis">
                    <div class="paso-titulo">
                        <span class="paso-numero">3</span>
                        Análisis
                    </div>
                    <div class="paso-texto">
                        Gemini interpreta los resultados y posibles causas.
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

        with flecha3:
            st.markdown(
                '<div class="flecha-proceso"><i class="bi bi-arrow-right"></i></div>',
                unsafe_allow_html=True
            )

        with paso4:
            st.markdown(
                """
                <div class="paso-analisis">
                    <div class="paso-titulo">
                        <span class="paso-numero">4</span>
                        Decisión
                    </div>
                    <div class="paso-texto">
                        Se generan recomendaciones para el supervisor.
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

        st.info("""
            💡 Selecciona el análisis individual para revisar un caso
            específico o el análisis histórico para identificar patrones.
            """)

        st.caption(
            "🔒 Los registros se utilizan únicamente para consulta. "
            "El análisis no modifica la información almacenada."
        )
        # =====================================================
        # SELECCIÓN DEL TIPO DE ANÁLISIS
        # =====================================================

        st.divider()

        st.markdown(
            '## <i class="bi bi-hand-index"></i> Selecciona el tipo de análisis',
            unsafe_allow_html=True
        )

        st.caption("Elige la información que deseas analizar mediante IA Generativa.")

        opcion1, opcion2 = st.columns(2)

        # -----------------------------------------------------
        # ÚLTIMA INSPECCIÓN
        # -----------------------------------------------------

        with opcion1:

            with st.container(border=True):
                st.caption("🟢 ANÁLISIS INDIVIDUAL")
                st.markdown(
                    '## <i class="bi bi-search"></i> Última inspección',
                    unsafe_allow_html=True
                )
                st.markdown(
                    """
                    <div style="font-size:20px; line-height:1.5;">
                        Analiza únicamente el registro más reciente almacenado
                        en VisionQA y genera un diagnóstico individual.
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                st.success("✔ Identificación de posibles causas raíz")
                st.success("✔ Acciones correctivas priorizadas")
                st.success("✔ Recomendaciones mediante metodología 6M")

                st.markdown(
                    '<div style="font-size: 17px; color: #6b7280;">'
                    '<i class="bi bi-clock"></i>&nbsp; Análisis individual del último registro disponible.'
                    '</div>',
                    unsafe_allow_html=True
                )

                if st.button(
                    "Analizar última inspección",
                    key="analizar_ultima_inspeccion",
                    width="stretch",
                ):

                    registros = obtener_inspecciones_api()

                    if not registros:

                        st.warning("No hay inspecciones registradas.")

                    else:

                        ultima = registros[0]

                        # -------- DATOS DE LAS TARJETAS --------

                        estado_tarjeta = (
                            str(ultima.get("resultado", "Sin datos")).strip().upper()
                        )

                        try:

                            confianza_tarjeta = float(ultima.get("confianza", 0))

                        except (TypeError, ValueError):

                            confianza_tarjeta = 0.0

                        defecto_tarjeta = (
                            str(ultima.get("defecto", "Sin defecto"))
                            .replace("_", " ")
                            .strip()
                            .title()
                        )

                        if not defecto_tarjeta:

                            defecto_tarjeta = "Sin defecto"

                        if estado_tarjeta in ["NO APTO", "MALA"]:

                            prioridad_tarjeta = "ALTA"
                            accion_tarjeta = "Contener pieza"

                        elif estado_tarjeta in ["APTO", "BUENA"]:

                            prioridad_tarjeta = "BAJA"
                            accion_tarjeta = "Liberar pieza"

                        else:

                            prioridad_tarjeta = "MEDIA"
                            accion_tarjeta = "Revisión manual"

                        # -------- DATOS PARA GEMINI --------

                        datos = f"""
Fecha: {ultima.get('fecha', 'Sin fecha')}
Resultado: {ultima.get('resultado', 'Sin resultado')}
Defecto: {ultima.get('defecto', 'Sin defecto')}
Confianza: {ultima.get('confianza', 0)}%
Origen: {ultima.get('origen', 'Sin origen')}
"""

                        with st.spinner(
                            "🧠 Gemini está analizando la última inspección..."
                        ):

                            resultado = analizar_causas(datos)

                        st.success("✅ Análisis completado correctamente.")

                        resultado_ia = resultado

                        titulo_informe = "🧠 Informe ejecutivo de la última inspección"

                        descripcion_informe = (
                            "Análisis basado en el registro más reciente "
                            "de VisionQA."
                        )

                        mostrar_tarjetas = True

        # -----------------------------------------------------
        # ÚLTIMAS 10 INSPECCIONES
        # -----------------------------------------------------

        with opcion2:

            with st.container(border=True):
                st.caption("🔵 ANÁLISIS HISTÓRICO")
                st.markdown(
                    '## <i class="bi bi-bar-chart-line"></i> Últimas 10 inspecciones',
                    unsafe_allow_html=True
                )

                st.markdown(
                    """
                    <div style="font-size:20px; line-height:1.5;">
                        Analiza el historial reciente para identificar defectos
                        recurrentes, tendencias y oportunidades de mejora.
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                st.info("✓ Identificación de tendencias")
                st.info("✓ Detección de defectos recurrentes")
                st.info("✓ Análisis comparativo del historial")

                st.markdown(
                    '<div style="font-size: 17px; color: #6b7280;">'
                    '<i class="bi bi-clock"></i>&nbsp; Análisis agrupado de hasta diez registros recientes.'
                    '</div>',
                    unsafe_allow_html=True
                )

                if st.button(
                    "Analizar historial",
                    key="analizar_ultimas_10_inspecciones",
                    width="stretch",
                ):

                    registros = obtener_inspecciones_api()

                    if not registros:

                        st.warning("No hay inspecciones registradas.")

                    else:

                        ultimas = registros[:10]

                        texto = ""

                        for registro in ultimas:

                            texto += f"""
Fecha: {registro.get('fecha', 'Sin fecha')}
Resultado: {registro.get('resultado', 'Sin resultado')}
Defecto: {registro.get('defecto', 'Sin defecto')}
Confianza: {registro.get('confianza', 0)}%
Origen: {registro.get('origen', 'Sin origen')}

"""

                        with st.spinner(
                            "🧠 Gemini está analizando las últimas inspecciones..."
                        ):

                            resultado = analizar_causas(texto)

                        st.success("✅ Análisis completado correctamente.")

                        resultado_ia = resultado

                        titulo_informe = (
                            "📊 Informe ejecutivo de las últimas " "10 inspecciones"
                        )

                        descripcion_informe = (
                            "Análisis de patrones, defectos y oportunidades "
                            "de mejora."
                        )

                        mostrar_tarjetas = False

    # =========================================================
    # MOSTRAR INFORME A TODO LO ANCHO
    # =========================================================

    if resultado_ia is not None:
        # =====================================================
        # INFORME EJECUTIVO
        # =====================================================

        st.markdown(
            """
            <div style="
                background:linear-gradient(135deg,#0F6CBD,#1593B7);
                border-radius:18px;
                padding:24px;
                color:white;
                margin-bottom:25px;
            ">

            <div style="
                font-size:30px;
                font-weight:700;
                margin-bottom:8px;
            ">
                <i class="bi bi-file-earmark-text"></i>&nbsp; Informe Ejecutivo IA
            </div>

            <div style="
                font-size:15px;
                opacity:0.95;
                line-height:1.6;
            ">
                Análisis generado automáticamente mediante Gemini AI a partir de
                los registros de inspección de VisionQA para apoyar la toma de
                decisiones en calidad.
            </div>

            </div>
            """,
            unsafe_allow_html=True,
        )

        if mostrar_tarjetas:

            (
                col_resultado,
                col_confianza,
                col_prioridad,
                col_defecto,
                col_accion,
            ) = st.columns(5)

            # -------- RESULTADO --------

            with col_resultado:

                with st.container(border=True):

                    if estado_tarjeta in ["APTO", "BUENA"]:
                        st.success("✓ Resultado")

                    elif estado_tarjeta in ["NO APTO", "MALA"]:
                        st.error("✕ Resultado")

                    else:
                        st.warning("⚠ Resultado")

                    st.metric(
                        label="Estado",
                        value=estado_tarjeta,
                    )

            # -------- CONFIANZA --------

            with col_confianza:

                with st.container(border=True):

                    st.info("▥ Confianza")

                    st.metric(
                        label="Nivel",
                        value=f"{confianza_tarjeta:.1f}%",
                    )

            # -------- PRIORIDAD --------

            with col_prioridad:

                with st.container(border=True):

                    if prioridad_tarjeta == "ALTA":
                        st.error("⚠ Prioridad")

                    elif prioridad_tarjeta == "MEDIA":
                        st.warning("⚠ Prioridad")

                    else:
                        st.success("✓ Prioridad")

                    st.metric(
                        label="Nivel",
                        value=prioridad_tarjeta,
                    )

            # -------- DEFECTO --------

            with col_defecto:

                with st.container(border=True):

                    st.warning("◎ Defecto")

                    st.metric(
                        label="Detectado",
                        value=defecto_tarjeta,
                    )

            # -------- ACCIÓN --------

            with col_accion:

                with st.container(border=True):

                    st.info("→ Acción")

                    st.metric(
                        label="Inmediata",
                        value=accion_tarjeta,
                    )

        # -------- INFORME GENERADO POR GEMINI --------

        resultado_limpio = (
            resultado_ia.replace("# 📄 Informe detallado", "")
            .replace("## 📄 Informe detallado", "")
            .replace("📄 Informe detallado", "")
            .replace(
                "Interpretación generada por IA utilizando la metodología "
                "Lean Manufacturing, Six Sigma y análisis 6M.",
                "",
            )
        )

        st.markdown(resultado_limpio.strip())
        # -------- GUARDAR INFORME PDF PARA REPORTES --------

        if mostrar_tarjetas:
            tipo_analisis_pdf = "Última inspección"
        else:
            tipo_analisis_pdf = "Últimas 10 inspecciones"

        pdf_informe = generar_pdf_informe_ia(
            resultado_ia=resultado_limpio,
            estado=estado_tarjeta,
            confianza=confianza_tarjeta,
            prioridad=prioridad_tarjeta,
            defecto=defecto_tarjeta,
            accion=accion_tarjeta,
            tipo_analisis=tipo_analisis_pdf,
        )

        nombre_pdf = (
            "VisionQA_Informe_IA_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".pdf"
        )

        st.session_state["pdf_informe_ia"] = pdf_informe
        st.session_state["nombre_pdf_ia"] = nombre_pdf
        st.session_state["fecha_informe_ia"] = datetime.now().strftime("%d/%m/%Y %H:%M")


def mostrar_reportes():
    st.markdown(
    """
    <style>

    /* Mensajes info, warning y error */
    div[data-testid="stAlert"] p {
        font-size: 20px !important;
    }

    /* Botones normales */
    div[data-testid="stButton"] button p {
        font-size: 20px !important;
        font-weight: 600 !important;
    }

    /* Botones de descarga */
    div[data-testid="stDownloadButton"] button p {
        font-size: 20px !important;
        font-weight: 600 !important;
    }

    </style>
    """,
    unsafe_allow_html=True
)
    st.markdown(
        '<span class="reportes-page-marker"></span>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '## <i class="bi bi-folder2-open"></i>&nbsp; Documentos disponibles',
        unsafe_allow_html=True
    )

    st.markdown(
        '<div style="font-size:18px; color:#6b7280;">Selecciona el documento que deseas generar o descargar.</div>',
        unsafe_allow_html=True
    )

    st.write("")

    col_excel, col_pdf = st.columns(2)

    # -------------------------------------------------
    # HISTORIAL
    # -------------------------------------------------

    with col_excel:

        with st.container(border=True):

            st.markdown(
                '<div style="font-size:30px;"><i class="bi bi-bar-chart-line"></i></div>',
                unsafe_allow_html=True
            )

            st.markdown("### Historial de inspecciones")

            st.markdown(
                """
                <div style="font-size:20px; line-height:1.5;">
                    Exporta todos los registros de inspección
                    almacenados por VisionQA en formato Excel.
                </div>
                """,
                unsafe_allow_html=True
            )
            st.info("Incluye fecha, resultado, confianza, " "defecto y origen.")

            registros = obtener_inspecciones_api()

            if not registros:

                st.warning("No existen inspecciones registradas para exportar.")

                st.button(
                    "📥 Excel no disponible",
                    key="excel_no_disponible",
                    width="stretch",
                    disabled=True,
                )

            else:

                try:

                    archivo_excel, total_registros = generar_excel_registros(registros)

                    nombre_excel = (
                        "Registro_VisionQA_"
                        + datetime.now().strftime("%Y%m%d_%H%M%S")
                        + ".xlsx"
                    )

                    st.download_button(
                        label="Descargar Excel",
                        data=archivo_excel,
                        file_name=nombre_excel,
                        mime=(
                            "application/vnd.openxmlformats-"
                            "officedocument.spreadsheetml.sheet"
                        ),
                        key="descargar_excel_reportes",
                        width="stretch",
                    )

                    st.markdown(
                        f'<div style="font-size:17px; color:#6b7280;">Registros disponibles: {total_registros}</div>',
                        unsafe_allow_html=True
                    )

                except Exception as error:

                    st.error("No fue posible preparar el archivo Excel.")

                    st.caption(str(error))
    # -------------------------------------------------
    # INFORME IA
    # -------------------------------------------------

    with col_pdf:

        with st.container(border=True):

            st.markdown(
                '<div style="font-size:30px;"><i class="bi bi-file-earmark-text"></i></div>',
                unsafe_allow_html=True
            )

            st.markdown("### Informe Ejecutivo IA")

            st.write("""
                Exporta el último análisis generado
                mediante Gemini IA en formato PDF.
                """)

            st.info("Incluye resumen, análisis 6M, " "acciones y recomendaciones.")

            if "pdf_informe_ia" in st.session_state:

                st.download_button(
                    label="Descargar PDF",
                    data=st.session_state["pdf_informe_ia"],
                    file_name=st.session_state.get(
                        "nombre_pdf_ia",
                        "VisionQA_Informe_IA.pdf",
                    ),
                    mime="application/pdf",
                    key="descargar_pdf_reportes",
                    width="stretch",
                )

                st.caption(
                    "Último informe generado: "
                    + st.session_state.get(
                        "fecha_informe_ia",
                        "Fecha no disponible",
                    )
                )

            else:

                st.warning("Primero genera un análisis en la sección IA Generativa.")

                st.button(
                    "PDF no disponible",
                    key="pdf_no_disponible",
                    width="stretch",
                    disabled=True,
                )

    st.divider()

    st.info("""
**Información**

• Los documentos se generan con la información almacenada en VisionQA.

• El historial de inspecciones se exporta en formato Excel (.xlsx).

• El informe ejecutivo se exporta en formato PDF (.pdf).
""")


def mostrar_footer():
    st.markdown(
        """
<style>
.visionqa-footer {
    margin-top: 24px;
    padding: 16px 4px 8px 4px;
    border-top: 1px solid #d9e1e8;
    color: #64748b;
    font-size: 14px;
    line-height: 1.5;
}
.visionqa-footer-title {
    color: #231F20;
    font-size: 15px;
    font-weight: 700;
    margin-bottom: 3px;
}
.visionqa-footer-info {
    margin-top: 10px;
}
.visionqa-footer-copy {
    margin-top: 10px;
    color: #94a3b8;
    font-size: 13px;
}
</style>
<div class="visionqa-footer">
<div class="visionqa-footer-title">VisionQA v1.0</div>
<div>Sistema Inteligente de Inspección Visual Asistido por Inteligencia Artificial</div>
<div class="visionqa-footer-info">Instituto Superior de Ciencias de Ciudad Juárez<br>Proyecto desarrollado para IOT Technologies</div>
<div class="visionqa-footer-copy">© 2026 VisionQA</div>
</div>
""",
        unsafe_allow_html=True,
    )


def mostrar_login():

    st.markdown(
        """
        <style>
        [data-testid="stSidebar"] {
            display: none;
        }

        [data-testid="stHeader"] {
            background: transparent;
        }

        .block-container {
            max-width: 1180px;
            padding-top: 3rem;
            padding-bottom: 3rem;
        }

        .login-panel {
            min-height: 620px;
            border-radius: 28px;
            overflow: hidden;
            box-shadow: 0 20px 55px rgba(35, 31, 32, 0.16);
            background: #FFFFFF;
            border: 1px solid rgba(189, 198, 195, 0.55);
        }

        .login-brand {
            min-height: 620px;
            padding: 58px 48px;
            border-radius: 28px 0 0 28px;
            background:
                radial-gradient(
                    circle at 15% 20%,
                    rgba(152, 218, 233, 0.38),
                    transparent 28%
                ),
                radial-gradient(
                    circle at 90% 82%,
                    rgba(255, 255, 255, 0.14),
                    transparent 30%
                ),
                linear-gradient(
                    145deg,
                    #0032A0 0%,
                    #1D7EAE 58%,
                    #1998B7 100%
                );
            color: white;
        }

        .iot-mark {
            font-size: 16px;
            font-weight: 700;
            letter-spacing: 1.4px;
            margin-bottom: 95px;
        }

        .brand-symbol {
            width: 72px;
            height: 72px;
            border: 3px solid #98DAE9;
            border-radius: 50%;
            position: relative;
            margin-bottom: 28px;
        }

        .brand-symbol::before,
        .brand-symbol::after {
            content: "";
            position: absolute;
            border: 3px solid #98DAE9;
            border-radius: 50%;
        }

        .brand-symbol::before {
            inset: 10px;
        }

        .brand-symbol::after {
            inset: 22px;
            background: #98DAE9;
        }

        .brand-title {
            font-size: 46px;
            line-height: 1;
            font-weight: 800;
            margin-bottom: 18px;
        }

        .brand-subtitle {
            font-size: 21px;
            line-height: 1.45;
            font-weight: 500;
            max-width: 420px;
            margin-bottom: 22px;
        }

        .brand-copy {
            font-size: 15px;
            line-height: 1.65;
            max-width: 440px;
            color: rgba(255, 255, 255, 0.82);
        }

        .brand-badge {
            display: inline-block;
            margin-top: 38px;
            padding: 10px 16px;
            border-radius: 999px;
            border: 1px solid rgba(255, 255, 255, 0.30);
            background: rgba(255, 255, 255, 0.10);
            font-size: 13px;
        }

        .login-form-header {
            margin-top: 85px;
            margin-bottom: 28px;
        }

        .login-form-header h1 {
            color: #231F20;
            font-size: 38px;
            margin: 0 0 10px 0;
        }

        .login-form-header p {
            color: #667085;
            font-size: 15px;
            margin: 0;
        }

        div[data-testid="stTextInput"] label {
            color: #231F20;
            font-weight: 600;
            font-size: 14px;
        }

        div[data-testid="stTextInput"] input {
            border-radius: 14px;
            min-height: 48px;
            border: 1px solid #BDC6C3;
            background: #FFFFFF;
        }

        div[data-testid="stTextInput"] input:focus {
            border-color: #1D7EAE;
            box-shadow: 0 0 0 3px rgba(29, 126, 174, 0.13);
        }

        div[data-testid="stButton"] > button {
            width: 100%;
            min-height: 50px;
            border: none;
            border-radius: 14px;
            background: linear-gradient(
                90deg,
                #0032A0 0%,
                #1D7EAE 100%
            );
            color: white;
            font-weight: 700;
            font-size: 15px;
            box-shadow: 0 8px 20px rgba(0, 50, 160, 0.22);
        }

        div[data-testid="stButton"] > button:hover {
            border: none;
            color: white;
            background: linear-gradient(
                90deg,
                #1D7EAE 0%,
                #0032A0 100%
            );
        }

        .login-help {
            margin-top: 24px;
            padding-top: 20px;
            border-top: 1px solid #E8ECEB;
            color: #667085;
            font-size: 13px;
            line-height: 1.5;
        }

        .login-footer {
            margin-top: 42px;
            color: #98A2B3;
            font-size: 12px;
        }
        div[data-testid="stHorizontalBlock"]
        > div[data-testid="column"]:first-child {
            background: linear-gradient(
                145deg,
                #0032A0 0%,
                #1D7EAE 58%,
                #1998B7 100%
            ) !important;
 
            min-height: 620px;
            padding: 58px 48px;
            border-radius: 28px 0 0 28px;
        }

        div[data-testid="stHorizontalBlock"]
        > div[data-testid="column"]:nth-child(2) {
            min-height: 620px;
            padding: 30px 42px;
            background: #FFFFFF;
            border-radius: 0 28px 28px 0;
        }

        @media (max-width: 900px) {
            .login-brand {
                min-height: 420px;
                border-radius: 28px 28px 0 0;
                padding: 42px 32px;
            }

            .login-form-header {
                margin-top: 20px;
            }

            .iot-mark {
                margin-bottom: 45px;
            }

            .brand-title {
                font-size: 38px;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    columna_marca, columna_formulario = st.columns([1.15, 0.85], gap="large")

    with columna_marca:

        st.markdown("<p class='iot-mark'>IOT TECHNOLOGIES</p>", unsafe_allow_html=True)

        st.markdown("<div class='brand-symbol'></div>", unsafe_allow_html=True)

        st.markdown("<h1 class='brand-title'>VisionQA</h1>", unsafe_allow_html=True)

        st.markdown(
            """
            <p class="brand-subtitle">
                Sistema Inteligente de Inspección Visual
            </p>
            """,
            unsafe_allow_html=True,
        )

        st.markdown(
            """
            <p class="brand-copy">
                Plataforma de control de calidad orientada a la
                detección de defectos, gestión de inspecciones y
                análisis de resultados para procesos industriales.
            </p>
            """,
            unsafe_allow_html=True,
        )

        st.markdown(
            """
            <div class="brand-badge">
                Conectividad · Confiabilidad · Eficiencia
            </div>
            """,
            unsafe_allow_html=True,
        )

    with columna_formulario:

        st.markdown(
            """
            <div class="login-form-header">
                <h1>Iniciar sesión</h1>
                <p>
                    Ingresa tus credenciales para acceder a VisionQA.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        correo = st.text_input(
            "Correo electrónico",
            placeholder="Escribe tu correo electrónico",
            key="login_correo",
        )

        contraseña = st.text_input(
            "Contraseña",
            type="password",
            placeholder="Escribe tu contraseña",
            key="login_contrasena",
        )

        iniciar = st.button("Iniciar sesión", key="login_boton")

        if iniciar:

            if not correo.strip() or not contraseña:
                st.warning("Escribe el correo electrónico y la contraseña.")

            else:

                try:

                    with st.spinner("Validando credenciales..."):

                        respuesta = requests.post(
                            "http://127.0.0.1:8000/api/login/",
                            json={"username": correo.strip(), "password": contraseña},
                            timeout=10,
                        )

                    if respuesta.status_code == 200:

                        datos = respuesta.json()

                        if datos.get("success"):

                            st.session_state["logueado"] = True
                            st.session_state["usuario"] = datos.get("username", correo)
                            st.session_state["is_staff"] = datos.get("is_staff", False)
                            st.session_state["is_superuser"] = datos.get(
                                "is_superuser", False
                            )

                            st.rerun()

                        else:

                            st.error("Usuario o contraseña incorrectos.")

                    else:

                        st.error("El servidor no pudo validar el acceso.")

                except requests.exceptions.ConnectionError:

                    st.error(
                        "No fue posible conectar con el servidor de "
                        "VisionQA. Verifica que Django esté ejecutándose."
                    )

                except requests.exceptions.Timeout:

                    st.error("El servidor tardó demasiado en responder.")

                except ValueError:

                    st.error("El servidor devolvió una respuesta no válida.")

                except Exception as error:

                    st.error(f"Ocurrió un error inesperado: {error}")

        st.markdown(
            """
            <div class="login-help">
                Acceso autorizado exclusivamente para personal
                registrado en el sistema.
            </div>

            <div class="login-footer">
                VisionQA v1.0 · IOT Technologies · 2026
            </div>
            """,
            unsafe_allow_html=True,
        )


def icono_svg(nombre, tamaño=22, margen_derecho=8):
    ruta = Path(__file__).parent / "assets" / "icons" / nombre

    if not ruta.exists():
        return ""

    contenido = base64.b64encode(ruta.read_bytes()).decode("utf-8")

    return (
        f"<img "
        f'src="data:image/svg+xml;base64,{contenido}" '
        f'width="{tamaño}" '
        f'height="{tamaño}" '
        f'style="'
        f"vertical-align:middle;"
        f"margin-right:{margen_derecho}px;"
        f"object-fit:contain;"
        f'">'
    )


def redirigir_a(url, mensaje="Redirigiendo..."):
    st.markdown(
        f"""
        <meta http-equiv="refresh" content="0; url={url}">
        <script>
            window.top.location.replace("{url}");
        </script>
        """,
        unsafe_allow_html=True,
    )

    st.info(mensaje)

    st.link_button(
        "Continuar",
        url,
        use_container_width=True,
    )

    st.stop()


@st.dialog("Centro de ayuda VisionQA")
def mostrar_ayuda():

    st.markdown("""
        ### ¿Cómo usar VisionQA?

        **1. Iniciar una inspección**  
        Ve al módulo **Inspección** y selecciona si deseas cargar una imagen o tomar una fotografía.

        **2. Interpretar el resultado**  
        - **APTO:** la pieza cumple con los criterios de calidad.
        - **NO APTO:** se detectó algún defecto en la pieza.

        **3. Consultar registros**  
        En el módulo **Registro** puedes revisar el historial de inspecciones.

        **4. Usar IA Generativa**  
        En el módulo **IA Generativa** puedes analizar causas, tendencias y recomendaciones.

        **5. Descargar reportes**  
        En **Reportes** puedes exportar el historial y los análisis generados.

        **6. Problemas con la cámara**  
        Verifica que Windows detecte la webcam y que el navegador tenga permiso para usarla.
        """)

    if st.button("Cerrar", key="cerrar_ayuda"):
        st.rerun()


def main():

    LOGIN_URL = "http://127.0.0.1:8000/api/login/"
    LOGOUT_URL = "http://127.0.0.1:8000/api/logout/"
    VERIFICAR_URL = "http://127.0.0.1:8000/api/verificar-acceso/"

    # ---------------------------------------------
    # CERRAR SESIÓN DESDE STREAMLIT
    # ---------------------------------------------

    if st.query_params.get("logout") == "1":

        st.session_state.clear()
        st.query_params.clear()

        redirigir_a(
            LOGOUT_URL,
            "Cerrando sesión...",
        )

    # ---------------------------------------------
    # OBTENER TOKEN
    # ---------------------------------------------

    token_url = st.query_params.get("token")

    if token_url:
        st.session_state["token_acceso"] = token_url

    token_acceso = st.session_state.get("token_acceso")

    # ---------------------------------------------
    # SI NO HAY TOKEN, VOLVER AL LOGIN
    # ---------------------------------------------

    if not token_acceso:

        st.session_state.clear()

        redirigir_a(
            LOGIN_URL,
            "Debes iniciar sesión para acceder a VisionQA.",
        )
    # ---------------------------------------------
    # VALIDAR SIEMPRE EL TOKEN CON DJANGO
    # ---------------------------------------------

    try:

        respuesta = requests.get(
            VERIFICAR_URL,
            params={
                "token": token_acceso,
            },
            timeout=10,
        )

        if respuesta.status_code != 200:

            st.session_state.clear()
            st.query_params.clear()

            redirigir_a(
                LOGIN_URL,
                "Tu sesión no es válida. Inicia sesión nuevamente.",
            )
        datos_usuario = respuesta.json()

        if not datos_usuario.get("autenticado"):

            st.session_state.clear()
            st.query_params.clear()

            redirigir_a(
                LOGIN_URL,
                "Tu sesión terminó.",
            )

        nombre = (
            datos_usuario.get("first_name")
            or datos_usuario.get("username")
            or "Usuario"
        )

        st.session_state["usuario"] = nombre
        st.session_state["username"] = datos_usuario.get(
            "username",
            "",
        )
        st.session_state["logueado"] = True

    except requests.RequestException:

        st.error("No fue posible conectar con el servidor de autenticación.")

        st.stop()

    # -------------------------------------------------
    # USUARIO AUTENTICADO
    # -------------------------------------------------

    nombre_usuario = st.session_state.get(
        "usuario",
        "Usuario",
    )
    if "tema" not in st.session_state:
        st.session_state["tema"] = "Claro"

    # -------- MENÚ LATERAL --------
    ahora = datetime.now()

    dias_semana = [
        "Lunes",
        "Martes",
        "Miércoles",
        "Jueves",
        "Viernes",
        "Sábado",
        "Domingo",
    ]

    meses = [
        "Ene",
        "Feb",
        "Mar",
        "Abr",
        "May",
        "Jun",
        "Jul",
        "Ago",
        "Sep",
        "Oct",
        "Nov",
        "Dic",
    ]

    dia_semana = dias_semana[ahora.weekday()]
    mes_actual = meses[ahora.month - 1]

    fecha_actual = f"{dia_semana}, {ahora.day:02d} " f"{mes_actual} {ahora.year}"

    hora_actual = ahora.strftime("%I:%M %p")

    if ahora.hour < 12:
        saludo = "Buenos días"
    elif ahora.hour < 19:
        saludo = "Buenas tardes"
    else:
        saludo = "Buenas noches"

    nombre_usuario = escape(
        str(
            st.session_state.get(
                "usuario",
                "Usuario",
            )
        )
    )
    if "mostrar_ayuda" not in st.session_state:
        st.session_state["mostrar_ayuda"] = False

    token_actual = st.query_params.get("token", "")
    html_barra = (
        '<div class="visionqa-topbar">'
        '<div class="visionqa-topbar-left">'
        f'<div class="visionqa-topbar-greeting">'
        f"{saludo}, {nombre_usuario}"
        "</div>"
        '<div class="visionqa-topbar-subtitle">'
        "Este es el estado actual del sistema VisionQA"
        "</div>"
        "</div>"
        '<div class="visionqa-topbar-right">'
        '<div class="visionqa-topbar-datetime">'
        f'<div class="visionqa-topbar-date">'
        f"{fecha_actual}"
        "</div>"
        f'<div class="visionqa-topbar-time">'
        f"{hora_actual}"
        "</div>"
        "</div>"
        '<div class="visionqa-topbar-actions">'
        f'<div class="visionqa-topbar-action" '
        f'title="{nombre_usuario}">'
        f"{icono_svg('user.svg',20,0)}"
        "</div>"
        f'<a class="visionqa-topbar-action" '
        f'href="?token={token_actual}&ayuda=1" '
        f'target="_self" '
        f'title="Ayuda">'
        f"{icono_svg('help.svg',20,0)}"
        "</a>"
        '<a class="visionqa-topbar-action" '
        'href="http://127.0.0.1:8501/?logout=1" '
        'target="_self" '
        'title="Cerrar sesión">'
        f"{icono_svg('logout.svg',20,0)}"
        "</a>"
        "</div>"
        "</div>"
        "</div>"
    )
    st.markdown(
        html_barra,
        unsafe_allow_html=True,
    )

    if st.query_params.get("ayuda") == "1":
        del st.query_params["ayuda"]
        mostrar_ayuda()
    # -------- RECORDAR PÁGINA ACTUAL --------

    opciones_menu = [
        "Dashboard",
        "Inspección",
        "Registro",
        "IA Generativa",
        "Reportes",
        "Acerca de",
    ]

    if "pagina_actual" not in st.session_state:
        st.session_state["pagina_actual"] = "Dashboard"

    indice_actual = opciones_menu.index(
        st.session_state["pagina_actual"]
    )
    with st.sidebar:

        st.image(
            "assets/logo_visionqa.png",
            width=220,
        )

        st.markdown(
            '<div class="sidebar-separator"></div>',
            unsafe_allow_html=True,
        )

        pagina = option_menu(
            menu_title=None,
            options=[
                "Dashboard",
                "Inspección",
                "Registro",
                "IA Generativa",
                "Reportes",
                "Acerca de",
            ],
            icons=[
                "speedometer2",
                "search",
                "clipboard-data",
                "cpu",
                "file-earmark-bar-graph",
                "info-circle",
            ],
            menu_icon=None,
            default_index=indice_actual,
            orientation="vertical",
            styles={
                "container": {
                    "padding": "16px 12px",
                    "margin": "0px",
                    "background-color": "#1998B7",
                    "border": "none",
                    "border-radius": "0px",
                    "box-shadow": "none",
                },
                "icon": {
                    "color": "#FFFFFF",
                    "font-size": "24px",
                },
                "nav-link": {
                    "font-size": "17px",
                    "font-weight": "600",
                    "color": "#FFFFFF",
                    "text-align": "left",
                    "margin": "8px 0",
                    "padding": "17px 18px",
                    "border-radius": "10px",
                    "background-color": "#1998B7",
                    "border": "none",
                },
                "nav-link-hover": {
                    "background-color": "rgba(255,255,255,0.16)",
                    "color": "#FFFFFF",
                },
                "nav-link-selected": {
                    "background-color": "#0A4E95",
                    "color": "#FFFFFF",
                    "font-weight": "700",
                    "border-radius": "10px",
                    "box-shadow": "0 4px 10px rgba(0,0,0,0.18)",
                },
            },
        )
        st.session_state["pagina_actual"] = pagina
        st.markdown(
            '<div class="sidebar-section">Apariencia</div>',
            unsafe_allow_html=True,
        )

        tema_seleccionado = st.radio(
            "Tema",
            ["Claro", "Oscuro"],
            horizontal=True,
            label_visibility="collapsed",
            index=(0 if st.session_state.get("tema", "Claro") == "Claro" else 1),
        )

        st.session_state["tema"] = tema_seleccionado

        st.markdown("---")

        st.markdown(
            """
                <div class="sidebar-footer">
                    <strong>VisionQA</strong><br>
                    Versión 1.0<br>
                    IOT Technologies
                </div>
                """,
            unsafe_allow_html=True,
        )

    # ---------------------------------------------
    # APLICAR TEMA FUERA DEL SIDEBAR
    # ---------------------------------------------

    if st.session_state.get("tema", "Claro") == "Oscuro":
        cargar_tema("oscuro.css")
    else:
        cargar_tema("claro.css")

    # -------- BARRA SUPERIOR DEL USUARIO --------

    # -------- DASHBOARD --------
    if pagina == "Dashboard":
        registros = obtener_inspecciones_api()

        total = len(registros)

        aptas = sum(
            1
            for registro in registros
            if str(registro.get("resultado", "")).strip().upper() == "APTO"
        )

        no_aptas = sum(
            1
            for registro in registros
            if str(registro.get("resultado", "")).strip().upper() == "NO APTO"
        )
        st.markdown(
            f"""
                <h2 style="
                    color:#231F20 !important;
                    font-size:40px;
                    font-weight:700;
                    margin-top:0;
                    margin-bottom:24px;
                    display:flex;
                    align-items:center;
                ">
                    {icono_svg("dashboard.svg", 48, 14)}
                    Dashboard Operativo
                </h2>
                """,
            unsafe_allow_html=True,
        )
        mostrar_resumen(total, aptas, no_aptas)

        mostrar_graficas(aptas, no_aptas)

        mostrar_indicadores(aptas, no_aptas)

    # -------- INSPECCIÓN --------

    elif pagina == "Inspección":

        mostrar_titulo(
            "inspection.svg",
            "Inspección Visual",
            "Captura y analiza piezas mediante inteligencia artificial.",
        )

        mostrar_estado_sistema()
        mostrar_modulo_inspeccion()

    # -------- REGISTRO --------

    elif pagina == "Registro":

        mostrar_titulo(
            "register.svg",
            "Registro de Inspecciones",
            "Consulta los resultados y el historial de inspecciones.",
        )

        mostrar_registro()
    # -------- IA GENERATIVA --------

    elif pagina == "IA Generativa":

        mostrar_titulo(
            "ai.svg",
            "Análisis Inteligente",
            "Analiza inspecciones mediante Gemini, metodología 6M, Lean Manufacturing y Six Sigma para apoyar la toma de decisiones en calidad.",
        )
        mostrar_gemini()

    # -------- REPORTES --------

    elif pagina == "Reportes":

        mostrar_titulo(
            "report.svg",
            "Centro de Reportes",
            "Genera y descarga los documentos disponibles del sistema VisionQA.",
        )

        mostrar_reportes()

    # -------- ACERCA DE --------

    elif pagina == "Acerca de":
        st.markdown(
            """
            <style>

            /* Texto normal de la sección Acerca de */
            div[data-testid="stMarkdownContainer"] p {
                font-size: 22px;
                line-height: 1.6;
            }

            /* Listas */
            div[data-testid="stMarkdownContainer"] li {
                font-size: 22px;
                line-height: 1.6;
            }

            </style>
            """,
            unsafe_allow_html=True
        )
        mostrar_titulo(
            "info.svg",
            "Acerca de VisionQA",
            "Información general del sistema y las tecnologías utilizadas.",
        )

        st.markdown("""
                **VisionQA** es un sistema inteligente de inspección visual
                desarrollado para apoyar el control de calidad de piezas
                manufacturadas.

                El sistema integra:

                - Visión por computadora.
                - Modelo de detección YOLOv8.
                - Procesamiento de imágenes con OpenCV.
                - Dashboard desarrollado con Streamlit.
                - Análisis de causas mediante Gemini.
                - Principios de Manufactura Esbelta y Six Sigma.
                """)

        st.markdown("### Información del proyecto")

        col1, col2 = st.columns(2)

        with col1:

            st.markdown("""
                    **Proyecto:** VisionQA

                    **Empresa:** IOT Technologies

                    **Área:** Control de Calidad
                    """)

        with col2:

            st.markdown("""
                    **Desarrolladora:** Dorcas Tabita Perez Martinez

                    **Tecnologías:** Python, YOLOv8, OpenCV, Streamlit y Gemini

                    **Versión:** 1.0
                    """)

        st.divider()

    mostrar_footer()


if __name__ == "__main__":
    main()
